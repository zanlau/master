import json
import requests
import urllib.parse
from openpyxl import load_workbook


wb = load_workbook('daten/population_ch.xlsx')
worksheet = wb['Bezirk']

municipality = []

for row in worksheet.iter_rows(min_row=2): #von erster Spalte über alle Elemente iterieren
    numbers = f"{row[0].value}, {row[1].value}, {row[2].value}"
    print(numbers)
    municipality.append({
        "number": row[0].value,
        "name": row[1].value,
        "population": row[2].value,
    })

data = {
    "source": "Bundesamt für Statistik - Bevölkerung"
              "https://www.pxweb.bfs.admin.ch/pxweb/de/px-x-0102010000_101/px-x-0102010000_101/px-x-0102010000_101.px",
    "data": municipality
}

with open('daten/population_ch_bezirk.json', 'w') as f:
    json.dump(data, f, indent=4)