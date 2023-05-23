import json
import requests
import urllib.parse
from openpyxl import load_workbook


wb = load_workbook('daten/air_ambulance_ch.xlsx')
worksheet = wb['Tabelle1']

stations = []

for row in worksheet.iter_rows(min_row=2): #von erster Spalte über alle Elemente iterieren
    address = f"{row[1].value}, {row[2].value}, {row[3].value}"
    print(address)
    url = f'https://nominatim.openstreetmap.org/search/{urllib.parse.quote(address)}?format=json'
    response = requests.get(url).json()
    stations.append({
        "lat": response[0]["lat"],
        "lon": response[0]["lon"],
        "institut": row[0].value,
        "stand_address": row[1].value,
        "stand_plz": row[2].value,
        "stand_ort": row[3].value,
    })

data = {
    "source": "Händische Erfassung mit Beizug der Literatur "
              "https://www.rega.ch/im-einsatz/rega-drohne",
    "data": stations
}

with open('daten/air_ambulance_ch.json', 'w') as f:
    json.dump(data, f, indent=4)