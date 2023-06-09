import json
import requests
import urllib.parse
from openpyxl import load_workbook


wb = load_workbook('daten/notfallstationen_ch.xlsx')
worksheet = wb['KZ2021_KZP21']

stations = []

for row in worksheet.iter_rows(min_row=2): #von erster Spalte über alle Elemente iterieren
    address = f"{row[3].value}, {row[4].value}, {row[5].value}"
    print(address)
    url = f'https://nominatim.openstreetmap.org/search/{urllib.parse.quote(address)}?format=json'
    response = requests.get(url).json()
    stations.append({
        "lat": response[0]["lat"],
        "lon": response[0]["lon"],
        "institut": row[2].value,
        "stand_address": row[3].value,
        "stand_plz": row[4].value,
        "stand_ort": row[5].value,
        "land": row[42].value,
        "personal_bestand": row[41].value,
    })

data = {
    "source": "BFS - Bundesamt für Statistik",
    "data": stations
}

with open('daten/notfallstationen_ch.json', 'w') as f:
    json.dump(data, f, indent=4)