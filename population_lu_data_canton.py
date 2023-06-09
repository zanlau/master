import json
import requests
import urllib.parse
from openpyxl import load_workbook


wb = load_workbook('daten/population_lu.xlsx')
worksheet = wb['Kanton']

municipality = []

for row in worksheet.iter_rows(min_row=2): #von erster Spalte über alle Elemente iterieren
    numbers = f"{row[0].value}, {row[1].value}"
    print(numbers)
    url = f'https://nominatim.openstreetmap.org/search/{urllib.parse.quote(numbers)}?format=json'
    response = requests.get(url).json()

    municipality.append({
        "lat": response[0]["lat"],
        "lon": response[0]["lon"],
        "number": row[0].value,
        "name": row[1].value,
        "population": row[2].value,
        "osm_id": response[0]["osm_id"]
    })

data = {
    "source": "Statec - Statistikseite des Staates Luxemburg"
              "https://lustat.statec.lu/vis?fs[0]=Topics%2C1%7CPopulation%20and%20employment%23B%23%7CPopulation%20structure%23B1%23&pg=0&fc=Topics&lc=en&df[ds]=ds-release&df[id]=DF_X021&df[ag]=LU1&df[vs]=1.0&pd=2015%2C2023&dq=A.",
    "data": municipality
}

with open('daten/population_lu_canton.json', 'w') as f:
    json.dump(data, f, indent=4)