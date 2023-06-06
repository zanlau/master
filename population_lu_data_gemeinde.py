import json
import requests
import urllib.parse
from openpyxl import load_workbook


wb = load_workbook('daten/population_lu.xlsx')
worksheet = wb['Gemeinde']

municipality = []

for row in worksheet.iter_rows(min_row=2): #von erster Spalte über alle Elemente iterieren
    numbers = f"{row[1].value}"
    print(numbers)
    url = f'https://nominatim.openstreetmap.org/search/{urllib.parse.quote(numbers)}?format=json'
    response = requests.get(url).json()
    print(response)

    for i, item in enumerate(response):
        url2 = f'https://nominatim.openstreetmap.org/lookup?osm_ids=R{item["osm_id"]}&format=json'
        response2 = requests.get(url2).json()[0]
        print(response2)
        address = response2["address"]
        if ("city" in address or "town" in address or "village" in address) and address["country_code"] == "lu":
            municipality.append({
                "lat": response[i]["lat"],
                "lon": response[i]["lon"],
                "number": row[0].value,
                "name": row[1].value,
                "population": row[2].value,
                "osm_id": response[i]["osm_id"]
            })
            break
        print("NEXT TRY")

data = {
    "source": "Statec - Statistikseite des Staates Luxemburg"
              "https://lustat.statec.lu/vis?fs[0]=Topics%2C1%7CPopulation%20and%20employment%23B%23%7CPopulation%20structure%23B1%23&pg=0&fc=Topics&lc=en&df[ds]=ds-release&df[id]=DF_X021&df[ag]=LU1&df[vs]=1.0&pd=2015%2C2023&dq=A.",
    "data": municipality
}

with open('daten/population_lu_gemeinde.json', 'w') as f:
    json.dump(data, f, indent=4)