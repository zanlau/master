import json
import base64
import requests
import urllib.parse
from openpyxl import load_workbook
from pathlib import Path


data_directory = "daten/population_ch_gemeinde"
Path(data_directory).mkdir(parents=True, exist_ok=True)


wb = load_workbook('daten/population_ch.xlsx')
worksheet = wb['Gemeinde']

municipality = []

def get_search_response(place):
    # use base64 encode to ensure the filename is valid
    filename = f"{data_directory}/{base64.urlsafe_b64encode(place.encode()).rstrip().decode()}.json"
    try:
        with open(filename, "r") as f:
            return json.load(f)
    except FileNotFoundError:
        url = f'https://nominatim.openstreetmap.org/search/{urllib.parse.quote(place)}?format=json&countrycodes=ch'
        response = requests.get(url).json()
        with open(filename, "w") as f:
            json.dump(response, f)
        return response

def get_lookup_response(osm_id):
    filename = f"{data_directory}/{osm_id}.json"
    try:
        with open(filename, "r") as f:
            return json.load(f)
    except FileNotFoundError:
        url = f'https://nominatim.openstreetmap.org/lookup?osm_ids={osm_id}&format=json&namedetails=1'
        response = requests.get(url).json()[0]
        with open(filename, "w") as f:
            json.dump(response, f)
        return response


for row in worksheet.iter_rows(min_row=2): #von erster Spalte über alle Elemente iterieren
    place = f"{row[1].value}".replace("/", " ")
    print(place)

    search_response = get_search_response(place)

    lookup_responses = []
    for item in search_response:
        osm_type = {"relation": "R", "node": "N", "way": "W"}[item["osm_type"]] #alle osm_types integrieren, da diese sonst nicht abgerufen werden können

        print(search_response)
        if (osm_type == "R" and item["class"] in ("boundary", "place")) or \
                (osm_type == "N" and item["class"] == "place" and item["type"] == "village"):
            lr = get_lookup_response(f'{osm_type}{item["osm_id"]}')
            if osm_type == "R":
                lookup_responses.insert(0, lr)
            else:
                lookup_responses.append(lr)

    for i, response in enumerate(lookup_responses):
        address = response["address"]
        other = False
        if village := address.get("village"):
            other |= village.replace("(", "").replace(")", "").replace("/", " ") == place
            if len(place) >= 3 and place[-3] == " ": # Abstand zwischen Ort und Kanton
                other |= place[:-3] == village

            if neighbourhood := address.get("neighbourhood"):
                other |= neighbourhood == place
            if place == address.get("suburb") and not address.get("town"):
                other |= True

            if response.get("namedetails", {}).get("official_name") == place:
                other |= True

            if "city_district" in address:
                print(response)
                other = False
        else:
            other = True

        matchers = ["city", "town", "village", "municipality", "hamlet", "suburb"]
        if any(x in address for x in matchers) and address["country_code"] == "ch" and other:
            municipality.append({
                "lat": response["lat"],
                "lon": response["lon"],
                "number": row[0].value,
                "name": row[1].value,
                "population": row[2].value,
                "osm_id": response["osm_id"]
            })
            break
        print("NEXT TRY")
    else:
        print("AAAAAAAAAA")
        raise "AAAA"


data = {
    "source": "Bundesamt für Statistik - Bevölkerung"
              "https://www.pxweb.bfs.admin.ch/pxweb/de/px-x-0102010000_101/px-x-0102010000_101/px-x-0102010000_101.px",
    "data": municipality
}

with open('daten/population_ch_gemeinde.json', 'w') as f:
    json.dump(data, f, indent=4)